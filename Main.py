from openpyxl import load_workbook as lw
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from Functions import setup
from datetime import date

merged = Workbook()
page = merged.active
excel = 1
count = 0

Spreadsheets, Directory = setup()
for books in Spreadsheets:
    direc = Directory + books
    print(direc)
    workbook = lw(direc)
    sheet = workbook.active
    column = 1
    row = 1
    a = True
    b = True
    while a:
        cell = get_column_letter(column) + "1"
        data = sheet[cell].value
        if data:
            column += 1
        else:
            break
    while b:
        cell = "A" + str(row)
        data = sheet[cell].value
        if data:
            row += 1
        else:
            break
    column -= 1
    row -= 1
    if excel == 1:
        minimum = 1
    else:
        minimum = 2
    for values in sheet.iter_rows(min_col=1, max_col=column, min_row=minimum, max_row=row, values_only=True):
        count += 1
        data = values
        lenght = len(values)
        for i in range(lenght):
            cell = get_column_letter(i + 1) + str(count)
            page[cell] = data[i]
    excel += 1
final = "./Merged/Merged" + str(date.today()) + ".xlsx"
merged.save(final)
